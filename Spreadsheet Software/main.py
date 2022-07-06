from pymongo import MongoClient
import pymongo
import openpyxl
from openpyxl import load_workbook
import time

warehouse_db = cluster["WAREHOUSE_MANAGEMENT"]
FINISHEDCollection = warehouse_db["FINISHED"]
Finished_JewlCollection = warehouse_db["FINISHED_JEWL"]

def gen(dnd,spreadsheet, collection, type):
    with open(dnd, "r") as collum:
        empty_collum = collum.readline()
    wb = load_workbook(spreadsheet)
    sheet = wb.active
    # date = sheet.cell(row = int(coa_collum), column = 5 # collom number is important)
    # date.value = str(today) # this is how we set data

    y = empty_collum
    x = 0

    document_list = ['Storage_Type', 'Date_Received', 'Store_Number','Contents', 'Date_Processed', 'MANIFEST_NUMBER', 'Problems']
    for document in collection.find():
        print(document)
        while x < 7:
            if x == 0:
                storage_type =  sheet.cell(row = int(empty_collum), column = 1)
                storage_type.value = str(document[document_list[x]])
            elif x == 1:
                Date_Received =  sheet.cell(row = int(empty_collum), column = 2)
                Date_Received.value = str(document[document_list[x]])
            elif x == 2:
                store_number = sheet.cell(row = int(empty_collum), column = 3)
                store_number.value = str(document[document_list[x]])
            elif x == 3:
                Contents = sheet.cell(row = int(empty_collum), column = 3)
                Contents.value = str(document[document_list[x]])
            elif x == 4:
                Date_Processed = sheet.cell(row = int(empty_collum), column = 3)
                Date_Processed.value = str(document[document_list[x]])
            elif x == 5:
                MANIFEST_NUMBER = sheet.cell(row = int(empty_collum), column = 3)
                MANIFEST_NUMBER.value = str(document[document_list[x]])
            elif x == 6:
                Problems = sheet.cell(row = int(empty_collum), column = 3)
                Problems.value = str(document[document_list[x]])
            else:
                nill = 'null'



    return
gen(dnd='DO_NOT_DELETE_SGW.txt',spreadsheet='Gay_lord_Toat_Log.xlsx', collection=FINISHEDCollection, type='gaylord_tote')
time.sleep(100)

while x < 17:
    a = json_object[x]
    list_of_dict_values = list(a.values())
    value = list_of_dict_values[0]
    if x == 16:
        devType = sheet.cell(row = int(coa_collum), column = 7)
        devType.value = str(value)
    elif x == 0:
        SN = sheet.cell(row = int(coa_collum), column = 8)
        SN.value = str(value)
    elif x == 1:
        brand = sheet.cell(row = int(coa_collum), column = 9)
        brand.value = str(value)
    elif x == 2:
        mod = sheet.cell(row = int(coa_collum), column = 10)
        mod.value = str(value)
    elif x == 3:
        pro = sheet.cell(row = int(coa_collum), column = 11)
        pro.value = str(value)
    elif x == 11:
        lic = sheet.cell(row = int(coa_collum), column = 14)
        lic.value = str(value)

        x+=1
    wb.save(filename='G:\My Drive\ECOM-COMPUTERS\Databases\FY21 MAR Citizenship COA report 2022 (April).xlsx')
    # Date:E Device Type:G SN:H Manufac:I model:J CPU:K COA:N
    print(y)
    with open('G:\My Drive\ECOM-COMPUTERS\Databases\DONTDELETE.txt', "w") as coa_coll:
        empty = int(y) + 1
        coa_coll.write(str(empty))
