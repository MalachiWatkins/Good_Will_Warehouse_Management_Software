from pymongo import MongoClient
import pymongo
import openpyxl
from openpyxl import load_workbook
import time
from plyer import notification


warehouse_db = cluster["WAREHOUSE_MANAGEMENT"]
FINISHEDCollection = warehouse_db["FINISHED"]
Finished_JewlCollection = warehouse_db["FINISHED_JEWL"]

def gen(dnd,spreadsheet, collection, type):
    with open(dnd, "r") as collum:
        empty_collum = collum.readline()
    wb = load_workbook(spreadsheet)
    sheet = wb.active

    # y = empty_collum
    x = 0

    document_list = ['Storage_Type', 'Date_Received', 'Store_Number','Contents', 'Date_Processed', 'MANIFEST_NUMBER', 'Problems']
    single_doc_list = []
    for document in collection.find():
        single_doc_list.append(document)
    if single_doc_list == []:
        notification.notify(
            title = 'No Processed Items',
            message = 'Nothing was processed right now!',
            timeout = 10,
        )
        exit()
    else:
        notification.notify(
            title = 'Adding Processed Itmes to Spreadsheet',
            message = 'The Process Has Started!',
            timeout = 5,
        )
        while x < len(single_doc_list):
            empty_collum = int(empty_collum) + 1
            print(single_doc_list[x]['Storage_Type'])
            storage_type =  sheet.cell(row = int(empty_collum), column = 1)
            storage_type.value = str(single_doc_list[x]['Storage_Type'])

            Date_Received =  sheet.cell(row = int(empty_collum), column = 2)
            Date_Received.value = str(single_doc_list[x]['Date_Received'])

            store_number =  sheet.cell(row = int(empty_collum), column = 3)
            store_number.value = str(single_doc_list[x]['Store_Number'])

            Contents =  sheet.cell(row = int(empty_collum), column = 4)
            Contents.value = str(single_doc_list[x]['Contents'])

            Date_Processed =  sheet.cell(row = int(empty_collum), column = 5)
            Date_Processed.value = str(single_doc_list[x]['Date_Processed'])

            MANIFEST_NUMBER =  sheet.cell(row = int(empty_collum), column = 6)
            MANIFEST_NUMBER.value = str(single_doc_list[x]['MANIFEST_NUMBER'])

            Problems =  sheet.cell(row = int(empty_collum), column = 7)
            Problems.value = str(single_doc_list[x]['Problems'])


            delquery = { "_id": float(single_doc_list[x]['_id']) }
            FINISHEDCollection.delete_one(delquery)

            x+=1
        with open('DO_NOT_DELETE_SGW.txt', "w") as coll:
            empty = empty_collum
            coll.write(str(empty))
        notification.notify(
            title = 'Finished Adding Items to Spreadsheet',
            message = 'All Done Here :)',
            timeout = 5,
        )
        wb.save(filename='Gay_lord_Toat_Log.xlsx')
    return

gen(dnd='DO_NOT_DELETE_SGW.txt',spreadsheet='Gay_lord_Toat_Log.xlsx', collection=FINISHEDCollection, type='gaylord_tote')
time.sleep(100)
